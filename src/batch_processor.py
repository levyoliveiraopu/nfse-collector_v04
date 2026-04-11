"""
Orquestrador do processamento em lote de NFS-e para todos os clientes.

Lê os clientes do CSV, busca os documentos fiscais na API ADN, gera os
relatórios Excel e faz upload no Google Drive, atualizando o estado NSU
a cada cliente processado com sucesso.
"""

import csv
import logging
import os
import time
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm

from src import auth, excel_builder, gdrive_uploader, nfse_fetcher, nsu_tracker

logger = logging.getLogger(__name__)

_COR_CABECALHO   = "1B5E20"
_COR_LINHA_PAR   = "F1F8E9"
_COR_LINHA_IMPAR = "FFFFFF"


# ---------------------------------------------------------------------------
# Configuração de logging
# ---------------------------------------------------------------------------

def _parse_bool_env(valor: str | None, default: bool = True) -> bool:
    """Converte valores textuais de ambiente para bool."""
    if valor is None:
        return default
    valor_norm = valor.strip().lower()
    if valor_norm in {"1", "true", "t", "yes", "y", "on"}:
        return True
    if valor_norm in {"0", "false", "f", "no", "n", "off"}:
        return False
    return default


def _parse_int_env(valor: str | None, default: int) -> int:
    """Converte variável de ambiente para int com fallback."""
    if valor is None:
        return default
    try:
        return int(valor.strip())
    except ValueError:
        return default


def _configurar_logging(log_level: str, log_to_console: bool = True) -> None:
    """Configura logging em arquivo e opcionalmente no console."""
    os.makedirs("logs", exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    log_path  = f"logs/execucao_{timestamp}.log"

    nivel = getattr(logging, log_level.upper(), logging.INFO)

    fmt = "%(asctime)s | %(levelname)-8s | %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"

    handlers: list[logging.Handler] = [
        logging.FileHandler(log_path, encoding="utf-8"),
    ]
    if log_to_console:
        handlers.append(logging.StreamHandler())

    logging.basicConfig(
        level=nivel,
        format=fmt,
        datefmt=datefmt,
        handlers=handlers,
    )
    logger.debug(
        "Log iniciado em '%s' (console=%s).",
        log_path,
        "on" if log_to_console else "off",
    )


# ---------------------------------------------------------------------------
# Leitura do CSV de clientes
# ---------------------------------------------------------------------------

def _carregar_clientes(csv_path: str, cnpj_filtro: str | None) -> list[dict]:
    """Lê config/clientes.csv e retorna a lista de clientes válidos.

    Valida que o arquivo de certificado existe antes de incluir o cliente.
    Se ``cnpj_filtro`` for informado, retorna apenas o cliente correspondente.

    Returns:
        Lista de dicts com chaves: cnpj, razao_social, cert_path, cert_password.

    Raises:
        FileNotFoundError: Se o arquivo CSV não existir.
        ValueError: Se o CNPJ filtrado não for encontrado no CSV.
    """
    if not os.path.isfile(csv_path):
        raise FileNotFoundError(f"Arquivo de clientes não encontrado: {csv_path}")

    clientes: list[dict] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for linha in reader:
            cnpj = linha.get("cnpj", "").strip()
            if not cnpj:
                continue

            if cnpj_filtro and cnpj != cnpj_filtro:
                continue

            cert_path = linha.get("cert_path", "").strip()
            if not os.path.isfile(cert_path):
                logger.warning(
                    "Certificado não encontrado para CNPJ %s ('%s'). Cliente ignorado.",
                    cnpj,
                    cert_path,
                )
                continue

            clientes.append({
                "cnpj":          cnpj,
                "razao_social":  linha.get("razao_social", "").strip(),
                "cert_path":     cert_path,
                "cert_password": linha.get("cert_password", "").strip(),
            })

    if cnpj_filtro and not clientes:
        raise ValueError(
            f"CNPJ '{cnpj_filtro}' não encontrado no CSV ou certificado ausente."
        )

    return clientes


# ---------------------------------------------------------------------------
# Excel consolidado
# ---------------------------------------------------------------------------

def _gerar_excel_consolidado(
    resultados: list[dict], ano: int, mes: int
) -> bytes:
    """Gera o Excel consolidado com uma linha por cliente processado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    cabecalhos = ["CNPJ", "Razão Social", "Qtd Notas", "Valor Total (R$)", "Status"]
    fill_cab  = PatternFill("solid", fgColor=_COR_CABECALHO)
    fill_par  = PatternFill("solid", fgColor=_COR_LINHA_PAR)
    fill_impar = PatternFill("solid", fgColor=_COR_LINHA_IMPAR)
    font_cab  = Font(bold=True, color="FFFFFF", size=11)
    alin_cen  = Alignment(horizontal="center", vertical="center")
    alin_dir  = Alignment(horizontal="right",  vertical="center")
    alin_esq  = Alignment(horizontal="left",   vertical="center")

    for col_idx, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = font_cab
        cell.fill = fill_cab
        cell.alignment = alin_cen
    ws.row_dimensions[1].height = 20

    for row_idx, res in enumerate(resultados, start=1):
        linha = row_idx + 1
        fill  = fill_par if row_idx % 2 == 0 else fill_impar

        valores = [
            res.get("cnpj"),
            res.get("razao_social"),
            res.get("qtd", 0),
            res.get("total", 0.0),
            res.get("status"),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=linha, column=col_idx, value=valor)
            cell.fill = fill
            if col_idx == 4:   # Valor Total
                cell.number_format = "#,##0.00"
                cell.alignment = alin_dir
            elif col_idx == 3:  # Qtd Notas
                cell.alignment = alin_cen
            else:
                cell.alignment = alin_esq

    from openpyxl.utils import get_column_letter
    larguras = [18, 42, 12, 22, 14]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# Processamento de um cliente
# ---------------------------------------------------------------------------

def _processar_cliente(
    cliente: dict,
    ano: int,
    mes: int,
    estado: dict,
    service,
    rate_limit_delay: int,
    max_documentos_por_execucao: int | None,
    nsu_estado_path: str,
    drive_root_id: str,
    dry_run: bool,
    todas_competencias: bool = False,
) -> dict:
    """Processa um único cliente. Retorna dict de resultado.

    Nunca lança exceção — erros são capturados e incluídos no resultado.
    """
    cnpj          = cliente["cnpj"]
    razao_social  = cliente["razao_social"]
    cert_path     = cliente["cert_path"]
    cert_password = cliente["cert_password"]

    cert_tmp: str | None = None
    key_tmp:  str | None = None

    try:
        logger.info("Processando %s (%s)", razao_social, cnpj)

        # b. Criar session mTLS
        session, cert_tmp, key_tmp, certificate = auth.criar_session_cliente(cert_path, cert_password)

        # c. Validar CNPJ do certificado (usa o objeto já carregado, sem reabrir o .pfx)
        cnpj_cert = auth.extrair_cnpj_do_certificado_obj(certificate, cert_path)
        if cnpj_cert != cnpj:
            logger.error(
                "CNPJ do certificado (%s) difere do CSV (%s). Cliente ignorado.",
                cnpj_cert,
                cnpj,
            )
            return {
                "cnpj": cnpj,
                "razao_social": razao_social,
                "status": "ERRO",
                "qtd": 0,
                "total": 0.0,
                "mensagem": (
                    f"CNPJ do certificado ({cnpj_cert}) "
                    f"difere do cadastrado no CSV ({cnpj})."
                ),
            }

        # d. Último NSU
        ultimo_nsu = nsu_tracker.obter_ultimo_nsu(estado, cnpj)

        # e. Buscar documentos
        lista_dfe, maior_nsu = nfse_fetcher.buscar_todos_dfe_novos(
            session,
            cnpj,
            ultimo_nsu,
            rate_limit_delay,
            max_documentos=max_documentos_por_execucao,
        )

        # e2. Filtrar apenas documentos do tipo NFSE
        lista_dfe = nfse_fetcher.filtrar_por_tipo_documento(lista_dfe)

        # f. Filtrar por competência (ou manter tudo, se solicitado)
        if todas_competencias:
            lista_mes = lista_dfe
        else:
            lista_mes = nfse_fetcher.filtrar_por_competencia(lista_dfe, ano, mes)

        # g. Sem notas no período
        if not lista_mes:
            if todas_competencias:
                logger.info("CNPJ %s: 0 notas (sem filtro de competência).", cnpj)
            else:
                logger.info("CNPJ %s: 0 notas em %02d/%d.", cnpj, mes, ano)
            if not dry_run and maior_nsu > ultimo_nsu:
                nsu_tracker.atualizar_nsu(estado, cnpj, maior_nsu)
                nsu_tracker.salvar_estado(estado, nsu_estado_path)
            return {
                "cnpj": cnpj,
                "razao_social": razao_social,
                "status": "ZERO_NOTAS",
                "qtd": 0,
                "total": 0.0,
            }

        # h. Extrair dados de cada nota e preparar XMLs para upload
        lista_dados: list[dict] = []
        lista_xmls: list[tuple[str, bytes]] = []

        for dfe in lista_mes:
            xml_content = nfse_fetcher.extrair_xml_do_doc(dfe)
            if not xml_content:
                logger.warning(
                    "Documento sem XML identificável — NSU=%s. Ignorado.",
                    dfe.get("NSU", "?"),
                )
                continue

            dados = nfse_fetcher.extrair_dados_nfse(xml_content)
            lista_dados.append(dados)

            # ChaveAcesso pode vir no envelope DistribuicaoNSU ou extraída do XML
            chave = (
                dfe.get("ChaveAcesso")
                or dados.get("chave_acesso")
                or str(dfe.get("NSU", "sem_chave"))
            )
            nome_xml = f"nfse_{chave}.xml"
            lista_xmls.append((nome_xml, xml_content.encode("utf-8")))

        # j. Gerar Excel
        excel_bytes = excel_builder.gerar_excel_cliente(
            lista_dados, cnpj, razao_social, ano, mes
        )

        # k. Upload no Drive
        if not dry_run:
            gdrive_uploader.organizar_e_enviar_cliente(
                service,
                drive_root_id,
                cnpj,
                razao_social,
                ano,
                mes,
                lista_xmls,
                excel_bytes,
            )

        # l. Atualizar e salvar NSU
        if not dry_run:
            nsu_tracker.atualizar_nsu(estado, cnpj, maior_nsu)
            nsu_tracker.salvar_estado(estado, nsu_estado_path)

        # n. Resultado
        total_servicos = sum(
            d.get("valor_servico") or 0.0
            for d in lista_dados
            if d.get("situacao") != "Cancelada"
        )

        logger.info(
            "CNPJ %s: %d nota(s) | R$ %.2f | NSU %d → %d%s",
            cnpj,
            len(lista_mes),
            total_servicos,
            ultimo_nsu,
            maior_nsu,
            " [DRY-RUN]" if dry_run else "",
        )

        return {
            "cnpj": cnpj,
            "razao_social": razao_social,
            "status": "OK",
            "qtd": len(lista_mes),
            "total": total_servicos,
        }

    except Exception as exc:  # noqa: BLE001
        logger.exception("Erro ao processar CNPJ %s: %s", cnpj, exc)
        return {
            "cnpj": cnpj,
            "razao_social": razao_social,
            "status": "ERRO",
            "qtd": 0,
            "total": 0.0,
            "mensagem": str(exc),
        }

    finally:
        # m. Limpar temporários sempre
        if cert_tmp or key_tmp:
            auth.limpar_temporarios(cert_tmp or "", key_tmp or "")


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def processar_todos_clientes(
    ano: int,
    mes: int,
    dry_run: bool = False,
    cnpj_filtro: str | None = None,
    lote: int | None = None,
    total_lotes: int | None = None,
    todas_competencias: bool = False,
) -> None:
    """Processa o lote completo de clientes para o período informado.

    Args:
        ano:          Ano de competência.
        mes:          Mês de competência (1–12).
        dry_run:      Se True, não faz uploads nem salva estado NSU.
        cnpj_filtro:  Se informado, processa apenas esse CNPJ.
        lote:         Número do lote a processar (1-based). Requer total_lotes.
        total_lotes:  Total de lotes em que os clientes serão divididos.
        todas_competencias: Se True, não filtra por mês/ano (usa todos os docs).
    """
    log_level      = os.getenv("LOG_LEVEL", "INFO")
    log_to_console = _parse_bool_env(os.getenv("LOG_TO_CONSOLE"), default=True)
    credentials    = os.getenv("GOOGLE_CREDENTIALS_JSON", "")
    drive_root_id  = os.getenv("GOOGLE_DRIVE_FOLDER_ROOT_ID", "")
    rate_limit_delay = int(os.getenv("RATE_LIMIT_DELAY", "3"))
    max_docs_env = _parse_int_env(os.getenv("MAX_DOCUMENTOS_POR_EXECUCAO"), 0)
    max_documentos_por_execucao = max_docs_env if max_docs_env > 0 else None
    nsu_estado_path  = os.getenv("NSU_ESTADO_PATH", "config/estado/ultimo_nsu.json")

    # 1. Logging
    _configurar_logging(log_level, log_to_console)

    inicio = time.time()
    periodo = f"{ano:04d}-{mes:02d}"

    # 2. Google Drive
    service = None
    if not dry_run:
        logger.info("Inicializando Google Drive...")
        service = gdrive_uploader.inicializar_drive(credentials)

    # 3. Clientes
    clientes = _carregar_clientes("config/clientes.csv", cnpj_filtro)

    # 3b. Fatiar por lote se solicitado
    if lote and total_lotes:
        import math
        tamanho_lote = math.ceil(len(clientes) / total_lotes)
        inicio = (lote - 1) * tamanho_lote
        fim = min(lote * tamanho_lote, len(clientes))
        logger.info(
            "Lote %d/%d: clientes %d–%d de %d total.",
            lote, total_lotes, inicio + 1, fim, len(clientes),
        )
        clientes = clientes[inicio:fim]

    logger.info("%d cliente(s) a processar.", len(clientes))
    if max_documentos_por_execucao is not None:
        logger.info(
            "Limite de busca por cliente nesta execução: %d documento(s).",
            max_documentos_por_execucao,
        )
    if todas_competencias:
        logger.info("Filtro de competência desativado (todas as competências).")

    # 4. Estado NSU
    estado = nsu_tracker.carregar_estado(nsu_estado_path)

    # 5. Processar cada cliente
    resultados: list[dict] = []

    for idx, cliente in enumerate(tqdm(clientes, desc="Clientes", unit="cliente", ncols=80)):
        resultado = _processar_cliente(
            cliente=cliente,
            ano=ano,
            mes=mes,
            estado=estado,
            service=service,
            rate_limit_delay=rate_limit_delay,
            max_documentos_por_execucao=max_documentos_por_execucao,
            nsu_estado_path=nsu_estado_path,
            drive_root_id=drive_root_id,
            dry_run=dry_run,
            todas_competencias=todas_competencias,
        )
        resultados.append(resultado)

        # o. Aguardar antes do próximo cliente
        if idx < len(clientes) - 1:
            time.sleep(rate_limit_delay)

    # 6. Excel consolidado
    nome_consolidado = f"NFS-e_CONSOLIDADO_{periodo}.xlsx"
    consolidado_bytes = _gerar_excel_consolidado(resultados, ano, mes)

    if not dry_run and service:
        try:
            gdrive_uploader.upload_ou_substituir(
                service,
                nome_arquivo=nome_consolidado,
                conteudo=consolidado_bytes,
                mimetype=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                pasta_id=drive_root_id,
            )
            logger.info("Consolidado enviado: '%s'.", nome_consolidado)
        except Exception as exc:  # noqa: BLE001
            logger.error("Falha ao enviar consolidado: %s", exc)

    # 7. Resumo final
    ok     = [r for r in resultados if r["status"] == "OK"]
    zeros  = [r for r in resultados if r["status"] == "ZERO_NOTAS"]
    erros  = [r for r in resultados if r["status"] == "ERRO"]
    minutos = (time.time() - inicio) / 60

    sep = "=" * 42
    print(f"\n{sep}")
    print(f"EXECUÇÃO CONCLUÍDA — {mes:02d}/{ano}")
    print(sep)
    print(f"✅ Processados com sucesso : {len(ok)} clientes")
    print(f"⚠️  Sem notas no período   : {len(zeros)} clientes")
    print(f"❌ Erros                   : {len(erros)} clientes")
    print("-" * 42)
    if erros:
        print("CNPJs com erro:")
        for r in erros:
            print(f"   {r['cnpj']} — {r.get('mensagem', 'sem detalhe')}")
    print(f"Tempo total: {minutos:.1f} minutos")
    print(f"{sep}\n")
