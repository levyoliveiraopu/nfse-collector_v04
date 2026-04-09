"""
Geração do relatório Excel (.xlsx) com as NFS-e mensais de um cliente.

Recebe a lista de dicts produzida por ``extrair_dados_nfse`` (nfse_fetcher.py),
formata os dados e retorna o conteúdo do arquivo como bytes (sem gravar em disco).

Abas geradas:
  - "Notas Emitidas": uma linha por nota, com colunas A–J.
  - "Resumo": totalizadores do período de competência.
"""

import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes de estilo
# ---------------------------------------------------------------------------

_COR_CABECALHO_FUNDO = "1B5E20"   # verde escuro
_COR_CABECALHO_TEXTO = "FFFFFF"   # branco
_COR_LINHA_PAR       = "F1F8E9"   # verde muito claro
_COR_LINHA_IMPAR     = "FFFFFF"   # branco
_COR_ATIVA           = "2E7D32"   # verde
_COR_CANCELADA       = "C62828"   # vermelho

_FILL_CABECALHO  = PatternFill("solid", fgColor=_COR_CABECALHO_FUNDO)
_FILL_LINHA_PAR  = PatternFill("solid", fgColor=_COR_LINHA_PAR)
_FILL_LINHA_IMPAR = PatternFill("solid", fgColor=_COR_LINHA_IMPAR)

_FONT_CABECALHO = Font(bold=True, color=_COR_CABECALHO_TEXTO, size=11)
_FONT_ATIVA     = Font(color=_COR_ATIVA)
_FONT_CANCELADA = Font(color=_COR_CANCELADA)
_FONT_CHAVE     = Font(size=8)

_ALINHAMENTO_CENTRO  = Alignment(horizontal="center", vertical="center")
_ALINHAMENTO_DIREITA = Alignment(horizontal="right",  vertical="center")
_ALINHAMENTO_ESQUERDA = Alignment(horizontal="left",  vertical="center")

_FORMATO_MOEDA = '#,##0.00'

# Definição das colunas da aba "Notas Emitidas": (cabeçalho, chave no dict)
_COLUNAS = [
    ("Número NFS-e",           "numero_nfse"),
    ("Data Emissão",           "data_emissao"),
    ("CNPJ/CPF Tomador",       "cnpj_tomador"),
    ("Razão Social Tomador",   "razao_social_tomador"),
    ("Descrição do Serviço",   "descricao_servico"),
    ("Valor do Serviço",       "valor_servico"),
    ("Valor ISS",              "valor_iss"),
    ("ISS Retido",             "iss_retido"),
    ("Situação",               "situacao"),
    ("Chave de Acesso",        "chave_acesso"),
]

# Índices das colunas (1-based) por papel funcional
_IDX_VALOR_SERVICO = 6   # coluna F
_IDX_VALOR_ISS     = 7   # coluna G
_IDX_ISS_RETIDO    = 8   # coluna H
_IDX_SITUACAO      = 9   # coluna I
_IDX_CHAVE         = 10  # coluna J


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def gerar_excel_cliente(
    lista_dados: list[dict],
    cnpj: str,
    razao_social: str,
    ano: int,
    mes: int,
) -> bytes:
    """Gera o relatório Excel com as NFS-e de um cliente no mês informado.

    Args:
        lista_dados:  Lista de dicts retornados por ``extrair_dados_nfse``.
        cnpj:         CNPJ de 14 dígitos do prestador (sem formatação).
        razao_social: Razão social do prestador.
        ano:          Ano de competência (ex.: 2025).
        mes:          Mês de competência (1–12).

    Returns:
        Conteúdo do arquivo ``.xlsx`` como ``bytes``.
        Nome sugerido: ``NFSe_{cnpj}_{ano:04d}-{mes:02d}.xlsx``.
    """
    wb = Workbook()

    aba_notas = wb.active
    aba_notas.title = "Notas Emitidas"

    aba_resumo = wb.create_sheet("Resumo")

    _preencher_notas(aba_notas, lista_dados)
    _preencher_resumo(aba_resumo, lista_dados, cnpj, razao_social, ano, mes)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    conteudo = buffer.read()

    logger.info(
        "Excel gerado — CNPJ=%s | %02d/%d | %d notas | %d bytes",
        cnpj,
        mes,
        ano,
        len(lista_dados),
        len(conteudo),
    )
    return conteudo


# ---------------------------------------------------------------------------
# Aba "Notas Emitidas"
# ---------------------------------------------------------------------------

def _preencher_notas(ws, lista_dados: list[dict]) -> None:
    """Preenche a aba 'Notas Emitidas' com cabeçalho e linhas de dados."""
    _escrever_cabecalho(ws)

    for idx, dados in enumerate(lista_dados, start=1):
        linha = idx + 1  # linha 1 = cabeçalho
        fill = _FILL_LINHA_PAR if idx % 2 == 0 else _FILL_LINHA_IMPAR
        _escrever_linha_nota(ws, linha, dados, fill)

    _ajustar_larguras(ws)
    ws.freeze_panes = "A2"


def _escrever_cabecalho(ws) -> None:
    """Escreve a linha de cabeçalho com estilo verde escuro."""
    for col_idx, (titulo, _) in enumerate(_COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = _FONT_CABECALHO
        cell.fill = _FILL_CABECALHO
        cell.alignment = _ALINHAMENTO_CENTRO

    ws.row_dimensions[1].height = 20


def _escrever_linha_nota(ws, linha: int, dados: dict, fill: PatternFill) -> None:
    """Escreve uma linha de dados de NFS-e com formatação adequada por coluna."""
    for col_idx, (_, chave) in enumerate(_COLUNAS, start=1):
        valor_bruto = dados.get(chave)
        cell = ws.cell(row=linha, column=col_idx)
        cell.fill = fill

        # --- Coluna ISS Retido ---
        if col_idx == _IDX_ISS_RETIDO:
            if valor_bruto is True:
                cell.value = "Sim"
            elif valor_bruto is False:
                cell.value = "Não"
            else:
                cell.value = None
            cell.alignment = _ALINHAMENTO_CENTRO
            continue

        # --- Colunas de valor monetário ---
        if col_idx in (_IDX_VALOR_SERVICO, _IDX_VALOR_ISS):
            cell.value = valor_bruto  # float ou None
            if valor_bruto is not None:
                cell.number_format = _FORMATO_MOEDA
            cell.alignment = _ALINHAMENTO_DIREITA
            continue

        # --- Coluna Situação ---
        if col_idx == _IDX_SITUACAO:
            cell.value = valor_bruto
            cell.alignment = _ALINHAMENTO_CENTRO
            if valor_bruto == "Cancelada":
                cell.font = _FONT_CANCELADA
            elif valor_bruto == "Ativa":
                cell.font = _FONT_ATIVA
            continue

        # --- Coluna Chave de Acesso ---
        if col_idx == _IDX_CHAVE:
            cell.value = valor_bruto
            cell.font = _FONT_CHAVE
            cell.alignment = _ALINHAMENTO_ESQUERDA
            continue

        # --- Demais colunas ---
        cell.value = valor_bruto
        cell.alignment = _ALINHAMENTO_ESQUERDA


def _ajustar_larguras(ws) -> None:
    """Ajusta a largura de cada coluna ao conteúdo, com limites por tipo."""
    for col_idx in range(1, len(_COLUNAS) + 1):
        col_letra = get_column_letter(col_idx)

        if col_idx == _IDX_CHAVE:
            ws.column_dimensions[col_letra].width = 50
            continue

        largura_max = max(
            (
                len(str(cell.value)) if cell.value is not None else 0
                for cell in ws[col_letra]
            ),
            default=0,
        )
        # Adiciona margem e limita a 40 caracteres
        ws.column_dimensions[col_letra].width = min(largura_max + 2, 40)


# ---------------------------------------------------------------------------
# Aba "Resumo"
# ---------------------------------------------------------------------------

def _preencher_resumo(
    ws,
    lista_dados: list[dict],
    cnpj: str,
    razao_social: str,
    ano: int,
    mes: int,
) -> None:
    """Preenche a aba 'Resumo' com totalizadores do período."""
    notas_ativas     = [d for d in lista_dados if d.get("situacao") != "Cancelada"]
    notas_canceladas = [d for d in lista_dados if d.get("situacao") == "Cancelada"]

    total_servicos = sum(
        d.get("valor_servico") or 0.0 for d in notas_ativas
    )
    total_iss = sum(
        d.get("valor_iss") or 0.0 for d in notas_ativas
    )
    iss_retido = sum(
        d.get("valor_iss") or 0.0
        for d in notas_ativas
        if d.get("iss_retido") is True
    )
    iss_a_recolher = total_iss - iss_retido

    # Estilo para rótulos (coluna A)
    fonte_rotulo = Font(bold=True)

    linhas = [
        ("Período de Competência",              f"{mes:02d}/{ano}"),
        ("CNPJ do Prestador",                   _formatar_cnpj(cnpj)),
        ("Razão Social",                        razao_social),
        (None,                                  None),
        ("Total de Notas Ativas",               len(notas_ativas)),
        ("Total de Notas Canceladas",           len(notas_canceladas)),
        (None,                                  None),
        ("Valor Total dos Serviços (ativas)",   total_servicos),
        ("Total ISS Retido pelo Tomador",       iss_retido),
        ("Total ISS a Recolher pelo Prestador", iss_a_recolher),
    ]

    for linha_idx, (rotulo, valor) in enumerate(linhas, start=1):
        if rotulo is None:
            continue

        celula_rotulo = ws.cell(row=linha_idx, column=1, value=rotulo)
        celula_rotulo.font = fonte_rotulo
        celula_rotulo.alignment = _ALINHAMENTO_ESQUERDA

        celula_valor = ws.cell(row=linha_idx, column=2, value=valor)
        celula_valor.alignment = _ALINHAMENTO_ESQUERDA

        # Formatar células monetárias
        if linha_idx >= 8:
            celula_valor.number_format = _FORMATO_MOEDA
            celula_valor.alignment = _ALINHAMENTO_DIREITA

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22


# ---------------------------------------------------------------------------
# Auxiliares
# ---------------------------------------------------------------------------

def _formatar_cnpj(cnpj: str) -> str:
    """Formata 14 dígitos como XX.XXX.XXX/XXXX-XX. Retorna inalterado se inválido."""
    digitos = "".join(c for c in (cnpj or "") if c.isdigit())
    if len(digitos) != 14:
        return cnpj or ""
    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
