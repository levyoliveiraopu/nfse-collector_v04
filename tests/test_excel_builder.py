"""Testes para o módulo excel_builder."""

from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from src import excel_builder
from worker_core.excel_builder import gerar_excel_periodo


def _carregar_wb(excel_bytes: bytes):
    return load_workbook(BytesIO(excel_bytes))


class TestGerarExcelCliente:
    _DADOS = [
        {
            "numero_nfse": "001",
            "data_emissao": "15/03/2025",
            "chave_acesso": "CHAVE001",
            "cnpj_prestador": "12345678000199",
            "razao_social_prestador": "Empresa ABC",
            "cnpj_tomador": "98765432000111",
            "razao_social_tomador": "Tomador XYZ",
            "descricao_servico": "Consultoria",
            "valor_servico": 1000.00,
            "valor_iss": 50.00,
            "iss_retido": False,
            "situacao": "Ativa",
        },
        {
            "numero_nfse": "002",
            "data_emissao": "20/03/2025",
            "chave_acesso": "CHAVE002",
            "cnpj_prestador": "12345678000199",
            "razao_social_prestador": "Empresa ABC",
            "cnpj_tomador": "11111111000100",
            "razao_social_tomador": "Outro Tomador",
            "descricao_servico": "Desenvolvimento",
            "valor_servico": 2000.00,
            "valor_iss": 100.00,
            "iss_retido": True,
            "situacao": "Ativa",
        },
    ]

    def test_retorna_bytes_validos(self):
        resultado = excel_builder.gerar_excel_cliente(
            self._DADOS, "12345678000199", "Empresa ABC", 2025, 3
        )
        assert isinstance(resultado, bytes)
        assert len(resultado) > 0

    def test_xlsx_tem_duas_abas(self):
        resultado = excel_builder.gerar_excel_cliente(
            self._DADOS, "12345678000199", "Empresa ABC", 2025, 3
        )
        wb = _carregar_wb(resultado)
        assert "Notas Emitidas" in wb.sheetnames
        assert "Resumo" in wb.sheetnames

    def test_aba_notas_tem_cabecalho_e_dados(self):
        resultado = excel_builder.gerar_excel_cliente(
            self._DADOS, "12345678000199", "Empresa ABC", 2025, 3
        )
        wb = _carregar_wb(resultado)
        ws = wb["Notas Emitidas"]
        # Cabeçalho na linha 1
        assert ws.cell(1, 1).value == "Número NFS-e"
        # Dados a partir da linha 2
        assert ws.cell(2, 1).value == "001"
        assert ws.cell(3, 1).value == "002"

    def test_aba_resumo_tem_totais(self):
        resultado = excel_builder.gerar_excel_cliente(
            self._DADOS, "12345678000199", "Empresa ABC", 2025, 3
        )
        wb = _carregar_wb(resultado)
        ws = wb["Resumo"]
        # Primeira linha: período
        assert ws.cell(1, 1).value == "Período de Competência"
        assert ws.cell(1, 2).value == "03/2025"
        # Total de notas ativas
        assert ws.cell(5, 2).value == 2

    def test_lista_vazia_gera_excel_valido(self):
        resultado = excel_builder.gerar_excel_cliente(
            [], "12345678000199", "Empresa ABC", 2025, 3
        )
        wb = _carregar_wb(resultado)
        assert "Notas Emitidas" in wb.sheetnames

    def test_nota_cancelada_nao_soma_no_total(self):
        dados = [
            {
                "numero_nfse": "001",
                "data_emissao": "15/03/2025",
                "chave_acesso": "C1",
                "cnpj_prestador": "12345678000199",
                "razao_social_prestador": "ABC",
                "cnpj_tomador": "99999999000100",
                "razao_social_tomador": "Tom",
                "descricao_servico": "Svc",
                "valor_servico": 500.00,
                "valor_iss": 25.00,
                "iss_retido": False,
                "situacao": "Cancelada",
            }
        ]
        resultado = excel_builder.gerar_excel_cliente(
            dados, "12345678000199", "ABC", 2025, 3
        )
        wb = _carregar_wb(resultado)
        ws = wb["Resumo"]
        # Total de notas ativas = 0
        assert ws.cell(5, 2).value == 0
        # Total de notas canceladas = 1
        assert ws.cell(6, 2).value == 1


def test_gerar_excel_periodo_exibe_intervalo_inclusivo() -> None:
    resultado = gerar_excel_periodo(
        [],
        "12345678000199",
        "Empresa ABC",
        date(2026, 6, 1),
        date(2026, 7, 31),
    )
    wb = _carregar_wb(resultado)
    assert wb["Resumo"].cell(1, 2).value == "01/06/2026 a 31/07/2026"
